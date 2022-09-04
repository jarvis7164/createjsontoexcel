# -*- coding: utf-8 -*-
# @Time    : 2022/8/27 22:53
# @Author  : Jarvis
# @Email   : jiamiao12@qq.com
# @File    : exporttoexcle.py
# @Software: PyCharm
import openpyxl as op

class write_excel_xlsx:

    def write_excel_xlsx(path, sheet_name, value):
        index = len(value)
        workbook = op.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
        workbook.save(path)
        print("xlsx格式表格写入数据成功！")


    def read_excel_xlsx(path, sheet_name):
        workbook = op.load_workbook(path)
        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        sheet = workbook[sheet_name]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()


